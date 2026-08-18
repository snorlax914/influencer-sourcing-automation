"""후보·판정·릴스를 한 장의 스프레드시트로 내보낸다.

대시보드는 세 CSV(candidates / verdicts / reels)를 화면에서만 합쳐 보여준다.
사람에게 넘기거나 구글 시트에 올릴 때는 그 합쳐진 표가 파일 하나로 필요하다.
`/api/candidates.csv`는 크롤 원본만 주므로 점수·등급·조회수가 빠져 있다.

xlsx로 쓰는 이유:
- 셀 타입이 있어 팔로워·점수가 **숫자로** 들어간다(CSV는 전부 문자열이라
  시트에서 정렬하면 "9157"이 "912"보다 앞에 오는 일이 생긴다).
- 소개글이 '='로 시작해도 수식으로 해석되지 않게 셀 단위로 못 박을 수 있다.
  인스타 소개글에는 '=' 같은 문자가 실제로 들어간다.
- 헤더 고정·필터·등급 색·프로필 링크를 파일 안에 넣어둘 수 있다.

구글 시트에서는 이 파일을 드라이브에 올려 바로 열면 되고, 서식도 유지된다.
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from discover import read_candidates
from history import follower_growth
from judge import load_verdicts
from reels import load_reels

# (헤더, 데이터 키, 열 너비, 종류). 종류는 셀 타입·서식을 정한다.
#   text  문자열 — 수식 해석을 막는다
#   int   숫자 — 천단위 구분 서식
#   link  프로필 URL — 하이퍼링크
#
# 순서는 '사람이 후보를 훑는 순서'다. 등급·점수를 맨 앞에 두고, 판단 근거는
# 뒤로 뺐다. 캡션은 길어서 맨 끝에 둔다 — 중간에 있으면 옆 열을 다 밀어낸다.
COLUMNS = [
    ("등급", "grade", 6, "text"),
    ("점수", "score", 7, "int"),
    ("계정", "username", 20, "text"),
    ("이름", "full_name", 18, "text"),
    ("프로필", "profile_url", 10, "link"),
    ("팔로워", "followers", 11, "int"),
    ("팔로워 증감", "follower_delta", 12, "int"),
    ("릴스 조회수(중앙값)", "views_median", 17, "int"),
    # 팔로워 수가 다른 계정끼리 도달을 비교할 수 있는 유일한 축이라 조회수 옆에 둔다.
    ("조회수 배율", "view_ratio", 11, "ratio"),
    ("릴스 표본", "reels_count", 9, "int"),
    ("티어", "tier", 7, "text"),
    ("지역 게이트", "gate", 10, "text"),
    ("선별", "screen", 7, "text"),
    ("게시물 수", "posts", 9, "int"),
    ("이메일", "email", 22, "text"),
    ("외부 링크", "external_url", 26, "text"),
    ("소개글", "biography", 40, "text"),
    ("채점 근거", "score_reason", 42, "text"),
    ("신호 근거", "reason", 34, "text"),
    ("선별 근거", "screen_reason", 34, "text"),
    ("릴스 메모", "reels_note", 18, "text"),
    ("최초 수집", "first_seen_at", 16, "text"),
    ("마지막 갱신", "updated_at", 16, "text"),
    ("최근 캡션", "captions", 60, "text"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
# 등급별 행 색. A·B는 컨택 대상이라 눈에 띄게, D는 눌러 놓는다.
GRADE_FILL = {
    "A": PatternFill("solid", fgColor="D1FAE5"),
    "B": PatternFill("solid", fgColor="ECFDF5"),
    "C": PatternFill("solid", fgColor="FEF9C3"),
    "D": PatternFill("solid", fgColor="F3F4F6"),
}

# 등급 없는 계정(선별 불합격·미판정)은 맨 뒤로. 등급 안에서는 점수 높은 순.
GRADE_ORDER = {"A": 0, "B": 1, "C": 2, "D": 3}


def _as_int(value):
    """CSV에서 온 문자열을 숫자로. 숫자가 아니면 원본을 그대로 돌려준다."""
    text = str(value if value is not None else "").strip().replace(",", "")
    return int(text) if text.lstrip("-").isdigit() else value


def _view_ratio(views_median, followers):
    """릴스 조회수 중앙값 / 팔로워. 둘 중 하나라도 없으면 빈 칸.

    스코어링(judge.py 도달 축)은 이 값을 티어별 구간으로 나눠 점수를 주지만,
    시트에서는 구간 이름보다 배율 자체가 쓸모 있다 — 같은 '높음' 안에서도
    1.6배와 2.9배는 다르고, 그 차이로 컨택 순서를 정하게 된다.
    """
    median, count = _as_int(views_median), _as_int(followers)
    if not isinstance(median, int) or not isinstance(count, int) or count <= 0 or median <= 0:
        return ""
    return round(median / count, 2)


def _short_time(value: str) -> str:
    """ISO 시각 → 'YYYY-MM-DD HH:MM'. 초·타임존은 표에서 읽을 일이 없다."""
    try:
        return datetime.fromisoformat(str(value).strip()).strftime("%Y-%m-%d %H:%M")
    except (ValueError, AttributeError):
        return str(value or "")


def export_rows() -> list[dict]:
    """세 CSV를 계정명으로 합쳐 내보내기용 평평한 행 목록을 만든다.

    webapp의 /api/candidates도 같은 세 파일을 합치지만, 그쪽은 화면용이라
    채점 표(breakdown)와 중첩 객체를 함께 싣는다. 스프레드시트는 한 칸에
    값 하나여야 하므로 여기서는 평평하게만 편다.
    """
    verdict_by_user = load_verdicts()
    reels_by_user = load_reels()
    growth = follower_growth()

    rows = []
    for row in read_candidates():
        username = row["username"]
        verdict = verdict_by_user.get(username, {})
        reel = reels_by_user.get(username, {})
        change = growth.get(username)
        rows.append(
            {
                **row,
                "followers": _as_int(row.get("followers")),
                "posts": _as_int(row.get("posts")),
                "first_seen_at": _short_time(row.get("first_seen_at")),
                "updated_at": _short_time(row.get("updated_at")),
                # 캡션 구분자(|||)는 시트에서 읽기 어려우니 줄바꿈으로 바꾼다.
                "captions": str(row.get("captions") or "").replace(" ||| ", "\n"),
                "screen": verdict.get("screen", ""),
                "screen_reason": verdict.get("screen_reason", ""),
                "score": verdict.get("score", ""),
                "grade": verdict.get("grade", ""),
                "tier": verdict.get("tier", ""),
                "gate": verdict.get("gate", ""),
                "score_reason": verdict.get("score_reason", ""),
                "reason": verdict.get("reason", ""),
                "views_median": reel.get("views_median", ""),
                "view_ratio": _view_ratio(reel.get("views_median"), row.get("followers")),
                "reels_count": reel.get("reels_count", ""),
                "reels_note": reel.get("note", ""),
                # 스냅샷이 2개 이상 쌓인 계정만 값이 있다(history.py 참고).
                "follower_delta": change["delta"] if change else "",
            }
        )

    rows.sort(
        key=lambda r: (
            GRADE_ORDER.get(r["grade"], 9),
            -(r["score"] if isinstance(r["score"], int) else 0),
            r["username"],
        )
    )
    return rows


def _write_cell(ws, row_idx: int, col_idx: int, value, kind: str):
    """셀 하나를 종류에 맞게 쓴다.

    text로 쓸 때 data_type을 못 박는 이유: openpyxl은 '='로 시작하는 문자열을
    수식으로 간주한다. 인스타 소개글에 '='가 들어오면 파일이 열릴 때
    #NAME? 오류가 되거나, 최악의 경우 의도치 않은 수식이 실행된다.
    값을 넣은 뒤 타입을 문자열로 되돌려 그냥 글자로 남게 한다.
    """
    cell = ws.cell(row=row_idx, column=col_idx)
    if kind == "int":
        cell.value = value if isinstance(value, int) else ""
        cell.number_format = "#,##0"
        return cell
    if kind == "ratio":
        # bool은 int의 하위형이라 isinstance(_, (int, float))에 걸린다 — 먼저 뺀다.
        cell.value = value if isinstance(value, (int, float)) and not isinstance(value, bool) else ""
        cell.number_format = '0.00"배"'
        return cell
    if kind == "link" and value:
        cell.value = "열기"
        cell.hyperlink = str(value)
        cell.font = Font(color="2563EB", underline="single")
        return cell
    cell.value = str(value if value is not None else "")
    if cell.data_type == "f":  # '='로 시작해 수식으로 잡힌 경우
        cell.data_type = "s"
    return cell


def build_workbook() -> bytes:
    """내보내기 xlsx를 메모리에서 만들어 바이트로 돌려준다."""
    rows = export_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "인플루언서 후보"

    for col_idx, (header, _key, width, _kind) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24

    for row_idx, row in enumerate(rows, start=2):
        fill = GRADE_FILL.get(row.get("grade"))
        for col_idx, (_header, key, _width, kind) in enumerate(COLUMNS, start=1):
            cell = _write_cell(ws, row_idx, col_idx, row.get(key, ""), kind)
            if fill:
                cell.fill = fill

    # 헤더 아래 고정 + 필터. 후보가 수백 줄이 되면 이게 없으면 훑기 어렵다.
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
